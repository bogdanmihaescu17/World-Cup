import csv
import io
import logging
import os
import re
from datetime import datetime, timedelta, timezone
from functools import wraps

import requests as http_requests

from dotenv import load_dotenv
from flask import (
    Flask,
    flash,
    redirect,
    render_template,
    request,
    send_file,
    url_for,
)
from flask_login import (
    LoginManager,
    UserMixin,
    current_user,
    login_required,
    login_user,
    logout_user,
)
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import check_password_hash, generate_password_hash
import openpyxl

load_dotenv()

app = Flask(__name__)
app.config["SECRET_KEY"] = os.getenv("SECRET_KEY", "dev-secret-change-me")
app.config["SQLALCHEMY_DATABASE_URI"] = os.getenv(
    "DATABASE_URL",
    "postgresql://postgres:postgres@localhost:5432/world_cup",
)
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = "login"


def configure_logging():
    level_name = os.getenv("LOG_LEVEL", "INFO").upper()
    level = getattr(logging, level_name, logging.INFO)
    app.logger.setLevel(level)
    if not app.logger.handlers:
        handler = logging.StreamHandler()
        handler.setLevel(level)
        formatter = logging.Formatter(
            "%(asctime)s %(levelname)s [%(name)s] %(message)s"
        )
        handler.setFormatter(formatter)
        app.logger.addHandler(handler)


configure_logging()


class User(UserMixin, db.Model):
    __tablename__ = "users"

    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    role = db.Column(db.String(20), nullable=False, default="user")
    created_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)

    predictions = db.relationship("Prediction", backref="user", lazy=True)

    def set_password(self, raw_password):
        self.password_hash = generate_password_hash(
            raw_password, method="pbkdf2:sha256"
        )

    def check_password(self, raw_password):
        return check_password_hash(self.password_hash, raw_password)


class Match(db.Model):
    __tablename__ = "matches"

    id = db.Column(db.Integer, primary_key=True)
    match_no = db.Column(db.Integer, unique=True, nullable=False)
    group_code = db.Column(db.String(16), nullable=True)
    team2_code = db.Column(db.String(16), nullable=True)
    team1 = db.Column(db.String(120), nullable=True)
    team2 = db.Column(db.String(120), nullable=True)
    kickoff_at = db.Column(db.DateTime, nullable=True)
    venue = db.Column(db.String(120), nullable=True)

    official_score1 = db.Column(db.Integer, nullable=True)
    official_score2 = db.Column(db.Integer, nullable=True)
    official_set_at = db.Column(db.DateTime, nullable=True)

    predictions = db.relationship("Prediction", backref="match", lazy=True)


class Prediction(db.Model):
    __tablename__ = "predictions"

    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("users.id"), nullable=False)
    match_id = db.Column(db.Integer, db.ForeignKey("matches.id"), nullable=False)
    pred_score1 = db.Column(db.Integer, nullable=False)
    pred_score2 = db.Column(db.Integer, nullable=False)
    created_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    updated_at = db.Column(
        db.DateTime, nullable=False, default=datetime.utcnow, onupdate=datetime.utcnow
    )

    __table_args__ = (db.UniqueConstraint("user_id", "match_id", name="uq_user_match"),)


class SpecialPrediction(db.Model):
    __tablename__ = "special_predictions"

    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("users.id"), unique=True, nullable=False)
    winner = db.Column(db.String(120), nullable=False, default="")
    goalscorer = db.Column(db.String(120), nullable=False, default="")
    created_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)

    user = db.relationship("User", backref=db.backref("special_prediction", uselist=False))


class OfficialSpecialResult(db.Model):
    __tablename__ = "official_special_results"

    id = db.Column(db.Integer, primary_key=True)
    winner = db.Column(db.String(120), nullable=False, default="")
    goalscorer = db.Column(db.String(120), nullable=False, default="")
    set_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)


@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))


def admin_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role != "admin":
            flash("Admin access required.", "error")
            return redirect(url_for("index"))
        return fn(*args, **kwargs)

    return wrapper


def match_outcome(score1, score2):
    if score1 == score2:
        return "draw"
    return "team1" if score1 > score2 else "team2"


def calculate_points(prediction, match):
    if match.official_score1 is None or match.official_score2 is None:
        return None
    if (
        prediction.pred_score1 == match.official_score1
        and prediction.pred_score2 == match.official_score2
    ):
        return 3
    if match_outcome(prediction.pred_score1, prediction.pred_score2) == match_outcome(
        match.official_score1, match.official_score2
    ):
        return 1
    return 0


PREDICTION_LOCK_HOURS = 2
SPECIAL_PREDICTION_WINNER_POINTS = 5
SPECIAL_PREDICTION_GOALSCORER_POINTS = 10

FOOTBALL_DATA_API_KEY = os.getenv("FOOTBALL_DATA_API_KEY", "")
FOOTBALL_DATA_BASE_URL = "https://api.football-data.org/v4"


def tournament_has_started():
    first = Match.query.filter(Match.kickoff_at.isnot(None)).order_by(
        Match.kickoff_at.asc()
    ).first()
    if not first:
        return False
    now = datetime.now(timezone.utc).replace(tzinfo=None)
    return now >= first.kickoff_at


def get_official_special():
    return OfficialSpecialResult.query.order_by(
        OfficialSpecialResult.id.desc()
    ).first()


def calculate_special_points(sp):
    official = get_official_special()
    if not official or (not official.winner and not official.goalscorer):
        return 0
    pts = 0
    if (
        official.winner
        and sp.winner
        and sp.winner.strip().lower() == official.winner.strip().lower()
    ):
        pts += SPECIAL_PREDICTION_WINNER_POINTS
    if (
        official.goalscorer
        and sp.goalscorer
        and sp.goalscorer.strip().lower() == official.goalscorer.strip().lower()
    ):
        pts += SPECIAL_PREDICTION_GOALSCORER_POINTS
    return pts


def prediction_is_locked(match):
    if not match.kickoff_at:
        return False
    now = datetime.now(timezone.utc).replace(tzinfo=None)
    return now >= match.kickoff_at - timedelta(hours=PREDICTION_LOCK_HOURS)


def can_submit_prediction(match, existing_prediction):
    if prediction_is_locked(match):
        return False
    if existing_prediction is not None:
        return False
    return True


_GROUP_CODE_LETTER = re.compile(r"^([A-L])(\d+)?$", re.IGNORECASE)


def group_letter_from_code(group_code):
    if not group_code:
        return None
    s = str(group_code).strip().upper()
    m = _GROUP_CODE_LETTER.match(s)
    if not m:
        return None
    return m.group(1)


def is_group_stage_match(match):
    return group_letter_from_code(match.group_code) is not None


def group_stage_matches_query():
    return [
        m
        for m in Match.query.order_by(
            Match.kickoff_at.asc().nullslast(), Match.match_no.asc()
        ).all()
        if is_group_stage_match(m)
    ]


def compute_group_standings_from_matches(matches):
    teams = {}
    for m in matches:
        for t in (m.team1, m.team2):
            if t not in teams:
                teams[t] = {
                    "team": t,
                    "pld": 0,
                    "w": 0,
                    "d": 0,
                    "l": 0,
                    "gf": 0,
                    "ga": 0,
                    "pts": 0,
                }
    for m in matches:
        if m.official_score1 is None or m.official_score2 is None:
            continue
        s1, s2 = m.official_score1, m.official_score2
        a, b = m.team1, m.team2
        teams[a]["pld"] += 1
        teams[b]["pld"] += 1
        teams[a]["gf"] += s1
        teams[a]["ga"] += s2
        teams[b]["gf"] += s2
        teams[b]["ga"] += s1
        if s1 > s2:
            teams[a]["w"] += 1
            teams[a]["pts"] += 3
            teams[b]["l"] += 1
        elif s2 > s1:
            teams[b]["w"] += 1
            teams[b]["pts"] += 3
            teams[a]["l"] += 1
        else:
            teams[a]["d"] += 1
            teams[b]["d"] += 1
            teams[a]["pts"] += 1
            teams[b]["pts"] += 1
    rows = []
    for t in teams.values():
        t["gd"] = t["gf"] - t["ga"]
        rows.append(t)
    rows.sort(
        key=lambda r: (-r["pts"], -r["gd"], -r["gf"], r["team"].lower()),
    )
    for idx, r in enumerate(rows, start=1):
        r["rank"] = idx
    return rows


def group_stage_ranking_rows(gs_ids=None):
    if gs_ids is None:
        gs_ids = {m.id for m in group_stage_matches_query()}
    users = User.query.order_by(User.username.asc()).all()
    rows = []
    for user in users:
        score = 0
        for p in user.predictions:
            if p.match_id not in gs_ids:
                continue
            pts = calculate_points(p, p.match)
            if pts is not None:
                score += pts
        rows.append({"username": user.username, "role": user.role, "score": score})
    rows.sort(key=lambda r: (-r["score"], r["username"].lower()))
    for idx, row in enumerate(rows, start=1):
        row["rank"] = idx
    return rows


_third_place_table_cache = None


def load_third_place_table():
    """Parse the AssignThird sheet from Excel. Cached after first load."""
    global _third_place_table_cache
    if _third_place_table_cache is not None:
        return _third_place_table_cache
    path = os.getenv("EXCEL_FILE_PATH", "World Cup_2026.xlsx")
    if not os.path.exists(path):
        _third_place_table_cache = {}
        return _third_place_table_cache
    wb = openpyxl.load_workbook(path, data_only=True)
    if "AssignThird" not in wb.sheetnames:
        _third_place_table_cache = {}
        return _third_place_table_cache
    ws = wb["AssignThird"]
    slot_codes = {}
    for col_idx in range(4, 12):
        val = ws.cell(4, col_idx).value
        if val:
            slot_codes[col_idx] = str(val).strip().upper()
    table = {}
    for row_idx in range(8, ws.max_row + 1):
        combo = ws.cell(row_idx, 3).value
        if not combo:
            continue
        combo = str(combo).strip().upper()
        row_map = {}
        for col_idx in range(4, 12):
            group_letter = ws.cell(row_idx, col_idx).value
            if group_letter and col_idx in slot_codes:
                row_map[slot_codes[col_idx]] = str(group_letter).strip().upper()
        table[combo] = row_map
    _third_place_table_cache = table
    return _third_place_table_cache


BRACKET_LAYOUT = [
    (74, "R32", 1, 1, 2), (77, "R32", 1, 3, 4),
    (73, "R32", 1, 5, 6), (75, "R32", 1, 7, 8),
    (83, "R32", 1, 9, 10), (84, "R32", 1, 11, 12),
    (81, "R32", 1, 13, 14), (82, "R32", 1, 15, 16),
    (76, "R32", 1, 17, 18), (78, "R32", 1, 19, 20),
    (79, "R32", 1, 21, 22), (80, "R32", 1, 23, 24),
    (86, "R32", 1, 25, 26), (88, "R32", 1, 27, 28),
    (85, "R32", 1, 29, 30), (87, "R32", 1, 31, 32),
    (89, "R16", 3, 1, 4), (90, "R16", 3, 5, 8),
    (93, "R16", 3, 9, 12), (94, "R16", 3, 13, 16),
    (91, "R16", 3, 17, 20), (92, "R16", 3, 21, 24),
    (95, "R16", 3, 25, 28), (96, "R16", 3, 29, 32),
    (97, "QF", 5, 1, 8), (98, "QF", 5, 9, 16),
    (99, "QF", 5, 17, 24), (100, "QF", 5, 25, 32),
    (101, "SF", 7, 1, 16), (102, "SF", 7, 17, 32),
    (104, "Final", 9, 1, 32),
]

BRACKET_CONNECTORS = [
    (2, 1, 4), (2, 5, 8), (2, 9, 12), (2, 13, 16),
    (2, 17, 20), (2, 21, 24), (2, 25, 28), (2, 29, 32),
    (4, 1, 8), (4, 9, 16), (4, 17, 24), (4, 25, 32),
    (6, 1, 16), (6, 17, 32),
    (8, 1, 32),
]


def compute_knockout_bracket():
    all_matches = Match.query.order_by(Match.match_no.asc()).all()
    match_map = {m.match_no: m for m in all_matches}

    gs_matches = [m for m in all_matches if is_group_stage_match(m)]
    groups = {}
    for m in gs_matches:
        letter = group_letter_from_code(m.group_code)
        if letter:
            groups.setdefault(letter, []).append(m)

    group_standings = {}
    group_complete = {}
    for letter, matches in sorted(groups.items()):
        group_standings[letter] = compute_group_standings_from_matches(matches)
        group_complete[letter] = all(
            m.official_score1 is not None and m.official_score2 is not None
            for m in matches
        )

    all_gs_played = all(group_complete.values()) if group_complete else False

    third_placed = []
    if all_gs_played:
        for letter, standings in group_standings.items():
            if len(standings) >= 3:
                team = standings[2]
                third_placed.append({
                    "group": letter,
                    "team": team["team"],
                    "pts": team["pts"],
                    "gd": team["gd"],
                    "gf": team["gf"],
                })
        third_placed.sort(key=lambda t: (-t["pts"], -t["gd"], -t["gf"], t["team"].lower()))
    qualifying_thirds = third_placed[:8]
    qualifying_groups = sorted([t["group"] for t in qualifying_thirds])
    combo_key = "".join(qualifying_groups)

    assign_table = load_third_place_table()
    slot_assignments = assign_table.get(combo_key, {})
    third_team_map = {t["group"]: t["team"] for t in qualifying_thirds}

    resolved_teams = {}

    def resolve_code(code):
        if not code:
            return None
        code = str(code).strip()
        m = re.match(r"^(\d)([A-L])$", code, re.IGNORECASE)
        if m:
            pos = int(m.group(1))
            group = m.group(2).upper()
            if not group_complete.get(group, False):
                return None
            standings = group_standings.get(group, [])
            if len(standings) >= pos:
                return standings[pos - 1]["team"]
            return None
        m = re.match(r"^3-([A-L]+)$", code, re.IGNORECASE)
        if m:
            if not all_gs_played:
                return None
            slot = code.upper()
            assigned_group = slot_assignments.get(slot)
            if assigned_group:
                return third_team_map.get(assigned_group)
            return None
        m = re.match(r"^W(\d+)$", code, re.IGNORECASE)
        if m:
            mn = int(m.group(1))
            if mn in resolved_teams:
                return resolved_teams[mn][2]
            return None
        m = re.match(r"^RU(\d+)$", code, re.IGNORECASE)
        if m:
            mn = int(m.group(1))
            if mn in resolved_teams:
                t1, t2, winner = resolved_teams[mn]
                if winner and winner == t1:
                    return t2
                elif winner and winner == t2:
                    return t1
            return None
        return None

    def get_winner(match_obj, team1_name, team2_name):
        if match_obj.official_score1 is None or match_obj.official_score2 is None:
            return None
        if match_obj.official_score1 > match_obj.official_score2:
            return team1_name
        elif match_obj.official_score2 > match_obj.official_score1:
            return team2_name
        return None

    knockout_matches = sorted(
        [m for m in all_matches if m.match_no >= 73],
        key=lambda m: m.match_no,
    )
    for ko_match in knockout_matches:
        mn = ko_match.match_no
        t1 = resolve_code(ko_match.group_code)
        t2 = resolve_code(ko_match.team2_code)
        if t1 is None and ko_match.team1:
            t1 = ko_match.team1
        if t2 is None and ko_match.team2:
            t2 = ko_match.team2
        winner = get_winner(ko_match, t1, t2)
        resolved_teams[mn] = (t1, t2, winner)

    bracket_entries = []
    for mn, round_name, col, row_start, row_end in BRACKET_LAYOUT:
        match = match_map.get(mn)
        if not match:
            continue
        t1, t2, winner = resolved_teams.get(mn, (None, None, None))
        bracket_entries.append({
            "match_no": mn,
            "round": round_name,
            "col": col,
            "row_start": row_start,
            "row_end": row_end,
            "team1": t1,
            "team2": t2,
            "team1_code": match.group_code or "",
            "team2_code": match.team2_code or "",
            "score1": match.official_score1,
            "score2": match.official_score2,
            "winner": winner,
            "venue": match.venue,
            "kickoff": match.kickoff_at,
        })

    third_match = match_map.get(103)
    third_place_entry = None
    if third_match:
        t1, t2, winner = resolved_teams.get(103, (None, None, None))
        third_place_entry = {
            "match_no": 103,
            "team1": t1,
            "team2": t2,
            "score1": third_match.official_score1,
            "score2": third_match.official_score2,
            "winner": winner,
        }

    final_winner = resolved_teams.get(104, (None, None, None))[2]

    return {
        "entries": bracket_entries,
        "connectors": BRACKET_CONNECTORS,
        "third_place": third_place_entry,
        "winner": final_winner,
        "all_gs_played": all_gs_played,
    }


def ensure_default_admin():
    admin_username = os.getenv("ADMIN_USERNAME", "admin")
    admin_password = os.getenv("ADMIN_PASSWORD", "admin123")
    admin = User.query.filter_by(username=admin_username).first()
    if not admin:
        admin = User(username=admin_username, role="admin")
        admin.set_password(admin_password)
        db.session.add(admin)
        db.session.commit()
    elif not admin.check_password(admin_password):
        admin.set_password(admin_password)
        db.session.commit()
        app.logger.info("Admin password re-hashed with current method.")


def import_excel_data(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    imported_matches = 0
    imported_users = 0

    if "Matches" in wb.sheetnames:
        ws = wb["Matches"]
        for row in ws.iter_rows(min_row=4, values_only=True):
            match_no = row[1]
            if not match_no or not isinstance(match_no, (int, float)):
                continue

            team1 = row[8]
            team2 = row[9]
            kickoff_at = row[5] if isinstance(row[5], datetime) else None
            if kickoff_at:
                kickoff_at = kickoff_at + timedelta(hours=1)
            group_code = str(row[2]).strip() if row[2] else None
            team2_code_val = str(row[3]).strip() if row[3] else None
            venue = str(row[7]).strip() if row[7] else None

            match = Match.query.filter_by(match_no=int(match_no)).first()
            if not match:
                match = Match(match_no=int(match_no))
                db.session.add(match)
                imported_matches += 1

            match.group_code = group_code
            match.team2_code = team2_code_val
            match.team1 = str(team1).strip() if team1 else None
            match.team2 = str(team2).strip() if team2 else None
            match.kickoff_at = kickoff_at
            match.venue = venue

    for ranking_sheet in ("Predictions_Ranking_1", "Predictions_Ranking_2"):
        if ranking_sheet not in wb.sheetnames:
            continue
        ws = wb[ranking_sheet]
        for row in ws.iter_rows(min_row=4, values_only=True):
            name = row[2] if len(row) > 2 else None
            if not name:
                continue
            username = str(name).strip()
            if not username:
                continue
            if not User.query.filter_by(username=username).first():
                u = User(username=username, role="user")
                u.set_password(os.getenv("DEFAULT_IMPORTED_USER_PASSWORD", "changeme123"))
                db.session.add(u)
                imported_users += 1

    db.session.commit()
    return imported_matches, imported_users


def auto_import_excel_if_empty():
    if os.getenv("AUTO_IMPORT_EXCEL", "true").lower() != "true":
        app.logger.info("Automatic Excel import disabled.")
        return

    if Match.query.count() > 0:
        app.logger.info("Skipping Excel auto-import because matches already exist.")
        return

    path = os.getenv("EXCEL_FILE_PATH", "World Cup_2026.xlsx")
    if not os.path.exists(path):
        app.logger.warning("Excel auto-import skipped; file not found at %s", path)
        return

    imported_matches, imported_users = import_excel_data(path)
    app.logger.info(
        "Excel auto-import completed. New matches=%s, new users=%s",
        imported_matches,
        imported_users,
    )


def ranking_rows():
    users = User.query.order_by(User.username.asc()).all()
    rows = []
    for user in users:
        score = 0
        for p in user.predictions:
            pts = calculate_points(p, p.match)
            if pts is not None:
                score += pts
        sp = SpecialPrediction.query.filter_by(user_id=user.id).first()
        if sp:
            score += calculate_special_points(sp)
        rows.append({"username": user.username, "role": user.role, "score": score})

    rows.sort(key=lambda r: (-r["score"], r["username"].lower()))
    for idx, row in enumerate(rows, start=1):
        row["rank"] = idx
    return rows


@app.route("/")
@login_required
def index():
    return render_template("index.html")


@app.route("/health")
def health():
    return {"status": "ok", "service": "world-cup-predictor"}, 200



@app.route("/register", methods=["GET", "POST"])
def register():
    if current_user.is_authenticated:
        return redirect(url_for("index"))
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        if len(username) < 3 or len(password) < 6:
            flash("Username min 3 chars, password min 6 chars.", "error")
            return redirect(url_for("register"))
        if User.query.filter_by(username=username).first():
            flash("Username already exists.", "error")
            return redirect(url_for("register"))
        user = User(username=username, role="user")
        user.set_password(password)
        db.session.add(user)
        db.session.commit()
        flash("Registration successful. Please login.", "success")
        return redirect(url_for("login"))
    return render_template("register.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user.is_authenticated:
        return redirect(url_for("index"))
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        user = User.query.filter_by(username=username).first()
        if not user or not user.check_password(password):
            flash("Invalid credentials.", "error")
            return redirect(url_for("login"))
        login_user(user)
        return redirect(url_for("index"))
    return render_template("login.html")


@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("login"))


@app.route("/predictions", methods=["GET", "POST"])
@login_required
def predictions():
    if request.method == "POST":
        match_id = int(request.form.get("match_id"))
        score1 = int(request.form.get("pred_score1"))
        score2 = int(request.form.get("pred_score2"))

        match = db.get_or_404(Match, match_id)
        existing = Prediction.query.filter_by(user_id=current_user.id, match_id=match.id).first()
        if not can_submit_prediction(match, existing):
            if existing:
                flash("You already submitted a prediction for this match. No changes allowed.", "error")
            else:
                flash("Predictions lock 2 hours before kickoff.", "error")
            return redirect(url_for("predictions"))

        pred = Prediction(user_id=current_user.id, match_id=match.id, pred_score1=score1, pred_score2=score2)
        db.session.add(pred)
        db.session.commit()
        flash("Prediction saved.", "success")
        return redirect(url_for("predictions"))

    matches = Match.query.order_by(Match.kickoff_at.asc().nullslast(), Match.match_no.asc()).all()
    predictions_map = {
        p.match_id: p
        for p in Prediction.query.filter_by(user_id=current_user.id).all()
    }
    return render_template(
        "predictions.html",
        matches=matches,
        predictions_map=predictions_map,
        calculate_points=calculate_points,
        can_submit_prediction=can_submit_prediction,
        prediction_is_locked=prediction_is_locked,
    )


@app.route("/special-prediction", methods=["GET", "POST"])
@login_required
def special_prediction():
    started = tournament_has_started()
    sp = SpecialPrediction.query.filter_by(user_id=current_user.id).first()
    official = get_official_special()

    if request.method == "POST":
        if started:
            flash("The tournament has started. Special predictions are locked.", "error")
            return redirect(url_for("special_prediction"))
        if sp:
            flash("You already submitted your special prediction. No changes allowed.", "error")
            return redirect(url_for("special_prediction"))

        winner = request.form.get("winner", "").strip()
        goalscorer = request.form.get("goalscorer", "").strip()
        if not winner or not goalscorer:
            flash("Both Winner and Goalscorer fields are required.", "error")
            return redirect(url_for("special_prediction"))

        sp = SpecialPrediction(
            user_id=current_user.id,
            winner=winner,
            goalscorer=goalscorer,
        )
        db.session.add(sp)
        db.session.commit()
        flash("Special prediction saved.", "success")
        return redirect(url_for("special_prediction"))

    sp_points = calculate_special_points(sp) if sp else 0
    return render_template(
        "special_prediction.html",
        sp=sp,
        started=started,
        official=official,
        sp_points=sp_points,
    )


@app.route("/group-stage")
@login_required
def group_stage():
    all_gs = group_stage_matches_query()
    letters = sorted(
        {group_letter_from_code(m.group_code) for m in all_gs},
    )
    groups_data = []
    for letter in letters:
        g_matches = [
            m for m in all_gs if group_letter_from_code(m.group_code) == letter
        ]
        groups_data.append(
            {
                "letter": letter,
                "standings": compute_group_standings_from_matches(g_matches),
            }
        )

    return render_template(
        "group_stage.html",
        groups_data=groups_data,
    )


@app.route("/knockout-stages")
@login_required
def knockout_stages():
    bracket = compute_knockout_bracket()
    return render_template("knockout_stages.html", bracket=bracket)


@app.route("/ranking")
@login_required
def ranking():
    return render_template("ranking.html", rows=ranking_rows())


@app.route("/admin")
@login_required
@admin_required
def admin_panel():
    matches = Match.query.order_by(Match.kickoff_at.asc().nullslast(), Match.match_no.asc()).all()
    users = User.query.order_by(User.username.asc()).all()
    official_special = get_official_special()
    return render_template("admin.html", matches=matches, users=users, official_special=official_special)


@app.route("/admin/users", methods=["POST"])
@login_required
@admin_required
def admin_create_user():
    username = request.form.get("username", "").strip()
    password = request.form.get("password", "").strip()
    role = request.form.get("role", "user").strip()
    if role not in ("admin", "user"):
        role = "user"
    if not username or len(password) < 6:
        flash("Username required, password min 6 chars.", "error")
        return redirect(url_for("admin_panel"))
    if User.query.filter_by(username=username).first():
        flash("Username already exists.", "error")
        return redirect(url_for("admin_panel"))
    user = User(username=username, role=role)
    user.set_password(password)
    db.session.add(user)
    db.session.commit()
    flash("User created.", "success")
    return redirect(url_for("admin_panel"))


@app.route("/admin/users/<int:user_id>/role", methods=["POST"])
@login_required
@admin_required
def admin_update_user_role(user_id):
    user = db.get_or_404(User, user_id)
    role = request.form.get("role", "user")
    if role not in ("admin", "user"):
        role = "user"
    user.role = role
    db.session.commit()
    flash("User role updated.", "success")
    return redirect(url_for("admin_panel"))


@app.route("/admin/users/<int:user_id>/delete", methods=["POST"])
@login_required
@admin_required
def admin_delete_user(user_id):
    user = db.get_or_404(User, user_id)
    if user.id == current_user.id:
        flash("You cannot delete your own admin account.", "error")
        return redirect(url_for("admin_panel"))
    Prediction.query.filter_by(user_id=user.id).delete()
    db.session.delete(user)
    db.session.commit()
    flash("User deleted.", "success")
    return redirect(url_for("admin_panel"))


@app.route("/admin/users/<int:user_id>/reset-password", methods=["POST"])
@login_required
@admin_required
def admin_reset_password(user_id):
    user = db.get_or_404(User, user_id)
    new_password = request.form.get("new_password", "").strip()
    if len(new_password) < 6:
        flash("Password must be at least 6 characters.", "error")
        return redirect(url_for("admin_panel"))
    user.set_password(new_password)
    db.session.commit()
    flash(f"Password reset for {user.username}.", "success")
    return redirect(url_for("admin_panel"))


@app.route("/admin/special-result", methods=["POST"])
@login_required
@admin_required
def admin_set_special_result():
    winner = request.form.get("winner", "").strip()
    goalscorer = request.form.get("goalscorer", "").strip()
    official = get_official_special()
    if not official:
        official = OfficialSpecialResult(winner=winner, goalscorer=goalscorer)
        db.session.add(official)
    else:
        official.winner = winner
        official.goalscorer = goalscorer
        official.set_at = datetime.utcnow()
    db.session.commit()
    flash("Official Winner / Goalscorer saved. Points recalculated.", "success")
    return redirect(url_for("admin_panel"))


@app.route("/admin/match-score", methods=["POST"])
@login_required
@admin_required
def admin_set_official_score():
    match_id = int(request.form.get("match_id"))
    score1 = int(request.form.get("official_score1"))
    score2 = int(request.form.get("official_score2"))
    match = db.get_or_404(Match, match_id)
    match.official_score1 = score1
    match.official_score2 = score2
    match.official_set_at = datetime.utcnow()
    db.session.commit()
    flash("Official score saved. Ranking is now updated.", "success")
    return redirect(url_for("admin_panel"))


@app.route("/admin/match-score/clear", methods=["POST"])
@login_required
@admin_required
def admin_clear_official_score():
    match_id = int(request.form.get("match_id"))
    match = db.get_or_404(Match, match_id)
    match.official_score1 = None
    match.official_score2 = None
    match.official_set_at = None
    db.session.commit()
    flash(f"Official score cleared for match #{match.match_no}.", "success")
    return redirect(url_for("admin_panel"))


def _normalize_team(name):
    """Normalize a team name for fuzzy matching between API and DB."""
    if not name:
        return ""
    name = name.strip().lower()
    replacements = {
        "korea republic": "south korea",
        "korea, republic of": "south korea",
        "ir iran": "iran",
        "côte d'ivoire": "ivory coast",
        "cote d'ivoire": "ivory coast",
        "united states": "usa",
    }
    for old, new in replacements.items():
        if old in name:
            name = new
            break
    return name


def sync_scores_from_api():
    """Fetch finished World Cup matches from football-data.org and update official scores."""
    api_key = FOOTBALL_DATA_API_KEY
    if not api_key:
        return 0, "FOOTBALL_DATA_API_KEY not configured."

    headers = {"X-Auth-Token": api_key}
    url = f"{FOOTBALL_DATA_BASE_URL}/competitions/WC/matches?status=FINISHED"

    try:
        resp = http_requests.get(url, headers=headers, timeout=15)
        resp.raise_for_status()
    except http_requests.RequestException as exc:
        return 0, f"API request failed: {exc}"

    data = resp.json()
    api_matches = data.get("matches", [])
    if not api_matches:
        return 0, "No finished matches returned by the API."

    db_matches = Match.query.all()
    team_index = {}
    for m in db_matches:
        key = (_normalize_team(m.team1), _normalize_team(m.team2))
        team_index[key] = m
        rev_key = (_normalize_team(m.team2), _normalize_team(m.team1))
        team_index[rev_key] = m

    updated = 0
    for api_m in api_matches:
        home_name = api_m.get("homeTeam", {}).get("name", "")
        away_name = api_m.get("awayTeam", {}).get("name", "")
        score_node = api_m.get("score", {})
        ft = score_node.get("fullTime", {})
        home_goals = ft.get("home")
        away_goals = ft.get("away")

        if home_goals is None or away_goals is None:
            continue

        key = (_normalize_team(home_name), _normalize_team(away_name))
        db_match = team_index.get(key)
        if not db_match:
            short_home = api_m.get("homeTeam", {}).get("shortName", "")
            short_away = api_m.get("awayTeam", {}).get("shortName", "")
            key2 = (_normalize_team(short_home), _normalize_team(short_away))
            db_match = team_index.get(key2)

        if not db_match:
            app.logger.warning("API match not found in DB: %s vs %s", home_name, away_name)
            continue

        is_home = _normalize_team(db_match.team1) == _normalize_team(home_name) or \
                  _normalize_team(db_match.team1) == _normalize_team(
                      api_m.get("homeTeam", {}).get("shortName", ""))

        if is_home:
            s1, s2 = home_goals, away_goals
        else:
            s1, s2 = away_goals, home_goals

        if db_match.official_score1 == s1 and db_match.official_score2 == s2:
            continue

        db_match.official_score1 = s1
        db_match.official_score2 = s2
        db_match.official_set_at = datetime.utcnow()
        updated += 1

    db.session.commit()
    return updated, None


@app.route("/admin/sync-api", methods=["POST"])
@login_required
@admin_required
def admin_sync_api():
    updated, error = sync_scores_from_api()
    if error:
        flash(f"API sync error: {error}", "error")
    elif updated == 0:
        flash("API sync complete. No new scores to update.", "success")
    else:
        flash(f"API sync complete. Updated {updated} match score(s).", "success")
    return redirect(url_for("admin_panel"))


@app.route("/admin/import-excel", methods=["POST"])
@login_required
@admin_required
def admin_import_excel():
    path = os.getenv("EXCEL_FILE_PATH", "World Cup_2026.xlsx")
    if not os.path.exists(path):
        flash(f"Excel file not found at: {path}", "error")
        return redirect(url_for("admin_panel"))
    matches, users = import_excel_data(path)
    flash(f"Excel imported. New matches: {matches}, new users: {users}.", "success")
    return redirect(url_for("admin_panel"))


@app.route("/admin/ranking-export")
@login_required
@admin_required
def admin_export_ranking():
    rows = ranking_rows()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["rank", "username", "role", "score"])
    for row in rows:
        writer.writerow([row["rank"], row["username"], row["role"], row["score"]])
    mem = io.BytesIO()
    mem.write(output.getvalue().encode("utf-8"))
    mem.seek(0)
    return send_file(
        mem,
        as_attachment=True,
        download_name="ranking.csv",
        mimetype="text/csv",
    )


@app.route("/admin/predictions")
@login_required
@admin_required
def admin_predictions():
    filter_user = request.args.get("user", "").strip()
    filter_group = request.args.get("group", "").strip()

    query = (
        db.session.query(Prediction, User, Match)
        .join(User, Prediction.user_id == User.id)
        .join(Match, Prediction.match_id == Match.id)
    )
    if filter_user:
        query = query.filter(User.username == filter_user)
    if filter_group:
        query = query.filter(Match.group_code.ilike(f"{filter_group}%"))
    query = query.order_by(User.username, Match.match_no)
    rows = query.all()

    prediction_rows = []
    for pred, user, match in rows:
        pts = calculate_points(pred, match)
        prediction_rows.append({
            "username": user.username,
            "match_no": match.match_no,
            "team1": match.team1 or match.group_code or "TBD",
            "team2": match.team2 or match.team2_code or "TBD",
            "pred_score1": pred.pred_score1,
            "pred_score2": pred.pred_score2,
            "official_score1": match.official_score1,
            "official_score2": match.official_score2,
            "points": pts,
            "submitted_at": pred.created_at,
        })

    special_preds = (
        db.session.query(SpecialPrediction, User)
        .join(User, SpecialPrediction.user_id == User.id)
        .order_by(User.username)
        .all()
    )
    if filter_user:
        special_preds = [
            (sp, u) for sp, u in special_preds if u.username == filter_user
        ]
    official = get_official_special()

    all_users = [u.username for u in User.query.order_by(User.username).all()]
    all_groups = sorted(
        {group_letter_from_code(m.group_code) for m in Match.query.all() if group_letter_from_code(m.group_code)}
    )

    return render_template(
        "admin_predictions.html",
        prediction_rows=prediction_rows,
        special_preds=special_preds,
        official=official,
        calculate_special_points=calculate_special_points,
        all_users=all_users,
        all_groups=all_groups,
        filter_user=filter_user,
        filter_group=filter_group,
    )


@app.route("/admin/predictions-export")
@login_required
@admin_required
def admin_export_predictions():
    rows = (
        db.session.query(Prediction, User, Match)
        .join(User, Prediction.user_id == User.id)
        .join(Match, Prediction.match_id == Match.id)
        .order_by(User.username, Match.match_no)
        .all()
    )
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "username", "match_no", "team1", "team2",
        "pred_score1", "pred_score2",
        "official_score1", "official_score2",
        "points", "submitted_at",
    ])
    for pred, user, match in rows:
        pts = calculate_points(pred, match)
        writer.writerow([
            user.username,
            match.match_no,
            match.team1 or match.group_code or "",
            match.team2 or match.team2_code or "",
            pred.pred_score1,
            pred.pred_score2,
            match.official_score1 if match.official_score1 is not None else "",
            match.official_score2 if match.official_score2 is not None else "",
            pts if pts is not None else "",
            pred.created_at,
        ])

    special_preds = (
        db.session.query(SpecialPrediction, User)
        .join(User, SpecialPrediction.user_id == User.id)
        .order_by(User.username)
        .all()
    )
    writer.writerow([])
    writer.writerow(["username", "winner", "goalscorer", "submitted_at"])
    for sp, user in special_preds:
        writer.writerow([user.username, sp.winner, sp.goalscorer, sp.created_at])

    mem = io.BytesIO()
    mem.write(output.getvalue().encode("utf-8"))
    mem.seek(0)
    return send_file(
        mem,
        as_attachment=True,
        download_name="predictions.csv",
        mimetype="text/csv",
    )


def _run_schema_migrations():
    """Add columns that db.create_all() won't create on existing tables."""
    from sqlalchemy import inspect, text

    inspector = inspect(db.engine)
    if "matches" in inspector.get_table_names():
        columns = {c["name"] for c in inspector.get_columns("matches")}
        if "team2_code" not in columns:
            db.session.execute(text("ALTER TABLE matches ADD COLUMN team2_code VARCHAR(16)"))
            app.logger.info("Migration: added matches.team2_code column.")
        if "team1" in columns:
            db.session.execute(text(
                "ALTER TABLE matches ALTER COLUMN team1 DROP NOT NULL"
            ))
            db.session.execute(text(
                "ALTER TABLE matches ALTER COLUMN team2 DROP NOT NULL"
            ))
    db.session.commit()


def init_database():
    with app.app_context():
        db.create_all()
        _run_schema_migrations()
        ensure_default_admin()
        auto_import_excel_if_empty()


if os.getenv("AUTO_INIT_DB", "true").lower() == "true":
    init_database()
    app.logger.info("Database auto-init completed.")
else:
    app.logger.info("Database auto-init skipped.")


if __name__ == "__main__":
    port = int(os.getenv("PORT", "5001"))
    debug = os.getenv("FLASK_DEBUG", "false").lower() == "true"
    app.logger.info("Starting app on port %s (debug=%s)", port, debug)
    app.run(host="0.0.0.0", port=port, debug=debug)
